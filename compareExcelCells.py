#! /usr/bin/python3
# compareExcelCells.py
# mauduong
# 22/04/2019
# Find duplicate cells between two excel spreadsheets.

import openpyxl, time

workBook1 = openpyxl.load_workbook('spreadsheet1.xlsx')
sheet1 = workBook1.active
workBook2 = openpyxl.load_workbook('spreadsheet2.xlsx')
sheet2 = workBook2.active

print('Searching...')
startTime = time.time()
numberOfDuplicates = 0

for (column, column2) in zip(sheet1.iter_cols(), sheet2.iter_cols()):
    for (cell, cell2) in zip(column, column2):
        if cell.value == cell2.value:
            numberOfDuplicates += 1
            print('Duplicate found.\nValue 1: ' + str(cell.value) + '\nValue 2: ' + str(cell2.value))
            print('Cell: ' + str(cell.column) + str(cell.row) + '\nColumn: ' + str(cell.column) + ' Row: ' + str(cell.row) + '\n')

endTime = time.time()
print('Time taken: ' + str(round((endTime - startTime) * 1000, 2)) + 'ms')
print('Duplicates found: ' + str(numberOfDuplicates))
