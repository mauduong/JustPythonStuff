#! /usr/bin/python3
# excelExample.py
# mauduong
# 21/04/2019

import openpyxl # sudo apt-get install python-openpyxl python3-openpyxl
workBook = openpyxl.load_workbook('example.xlsx')

print(workBook)

sheet = workBook['Sheet1']
print(sheet)
print(type(sheet))
print(sheet.title)
print(sheet['A1'])
print(sheet['A1'].value)
print(sheet['B2'].value)

print('Row: ' + str(sheet['B2'].row) + '\nColumn: ' + sheet['B2'].column + '\nValue: ' + sheet['B2'].value)

getActiveSheet = workBook.active
print(getActiveSheet)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=2, column=1).value)

for cell in range(1, 8, 2):
    print(cell, sheet.cell(row=cell, column=2).value)
